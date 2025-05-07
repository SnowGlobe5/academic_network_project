import os
import json
import re
from openpyxl import Workbook
from datetime import datetime

# Folder containing the experiment folders
base_folder = "/data/sabrina/academic_network_project/anp_models"

date_lower = "2024_11_08"
date_upper = "2027_04_20"
lr_lower = 0.00001
lr_upper = 0.01

# Convert date strings to integers for comparison
date_lower_int = int(date_lower.replace("_", ""))
date_upper_int = int(date_upper.replace("_", ""))

# Regex for folder names
folder_pattern = re.compile(r'anp_link_prediction_co_author_(.*?)_(\d{4}_\d{2}_\d{2})_\d{2}_\d{2}_\d{2}')

# Dictionaries to store results for normal and improved
results_normal = {}
results_improved = {}

def read_info_json(folder):
    info_file = os.path.join(folder, "info.json")
    if os.path.exists(info_file):
        with open(info_file) as f:
            return json.load(f)
    return None

def read_log_json(folder):
    for file_name in os.listdir(folder):
        if file_name.endswith("_log.json"):
            log_file = os.path.join(folder, file_name)
            with open(log_file) as f:
                return json.load(f)
    return None

# Traverse experiment folders
for experiment_folder in os.listdir(base_folder):
    experiment_path = os.path.join(base_folder, experiment_folder)
    if not os.path.isdir(experiment_path):
        continue

    match = folder_pattern.match(experiment_folder)
    if not match:
        continue

    folder_date = match.group(2)
    if not (date_lower_int <= int(folder_date.replace("_", "")) <= date_upper_int):
        continue

    info = read_info_json(experiment_path)
    log_data = read_log_json(experiment_path)
    if not info or not log_data:
        continue

    lr = info.get("lr", None)
    if lr is None or not (lr_lower <= float(lr) <= lr_upper):
        continue

    # Get classification: improved or not
    is_improved = "improved_model" in experiment_folder
    results = results_improved if is_improved else results_normal

    gt_type = info.get("GT_infosphere_type", "N/A")

    # Prepare metrics
    val_acc_list = log_data.get("validation_accuracy_list", [])
    loss_list = log_data.get("validation_loss_list", [])

    val_acc = val_acc_list[-1] if val_acc_list else "N/A"
    highest_acc = max(val_acc_list) if val_acc_list else "N/A"
    avg_last10_acc = sum(val_acc_list[-10:]) / 10 if len(val_acc_list) >= 10 else "N/A"

    last_loss = loss_list[-1] if loss_list else "N/A"
    min_loss = min(loss_list) if loss_list else "N/A"
    avg_last10_loss = sum(loss_list[-10:]) / 10 if len(loss_list) >= 10 else "N/A"

    # Flatten all info.json entries
    flat_info = {k: str(v) for k, v in info.items()}

    row = [
        experiment_folder,
        lr,
        val_acc,
        highest_acc,
        avg_last10_acc,
        last_loss,
        min_loss,
        avg_last10_loss,
        len(val_acc_list)
    ] + [flat_info.get(k, "N/A") for k in sorted(flat_info.keys())]

    if gt_type not in results:
        results[gt_type] = {"data": [], "columns": set(sorted(flat_info.keys()))}
    else:
        results[gt_type]["columns"].update(flat_info.keys())

    results[gt_type]["data"].append(row)

def write_results_to_excel(results, filename_prefix):
    wb = Workbook()
    for gt_type, content in results.items():
        ws = wb.create_sheet(title=f"GT_{gt_type}")

        # Headers
        info_keys = sorted(content["columns"])
        headers = [
            "Folder", "Learning Rate", "Validation Accuracy", "Highest Accuracy", "Avg Last 10 Acc",
            "Last Loss", "Min Loss", "Avg Last 10 Loss", "Epochs"
        ] + info_keys
        ws.append(headers)

        # Add rows
        for row in content["data"]:
            # Fill missing keys
            full_row = row[:9] + [row[9 + info_keys.index(k)] if k in info_keys else "N/A" for k in info_keys]
            ws.append(full_row)

    # Remove default sheet if exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Save the workbook
    now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    wb.save(f"{filename_prefix}_{now_str}.xlsx")

# Write both files
write_results_to_excel(results_normal, "report_normal")
write_results_to_excel(results_improved, "report_improved")

print("Excel reports created successfully.")
