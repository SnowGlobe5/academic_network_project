import os
import json
import re
from openpyxl import Workbook
from datetime import datetime

# Folder containing the experiment folders
base_folder = "/home/sabrina/academic_network_project/anp_models"

date_lower = "2024_11_08"
date_upper = "2024_11_10"
lr_lower = 0.00001
lr_upper = 0.01

infosphere_string = ["no_infosph", "future", "top_paper", "top_paper_topic", "N", "L"]
# Convert date strings to integers for comparison
date_lower_int = int(date_lower.replace("_", ""))
date_upper_int = int(date_upper.replace("_", ""))

# Dictionary to organize the results
results = [[{}, {}, {}, {}, {}, {}], [{}, {}, {}, {}, {}, {}]]

# Function to read the info.json file
def read_info_json(folder):
    info_file = os.path.join(folder, "info.json")
    if os.path.exists(info_file):
        with open(info_file) as f:
            info_data = json.load(f)
            if info_data['infosphere_type'] == 0:
                del info_data['infosphere_parameters']
            return info_data
    return None

# Function to read the *_log.json file
def read_log_json(folder):
    log_file = None
    for file_name in os.listdir(folder):
        if file_name.endswith("_log.json"):
            log_file = os.path.join(folder, file_name)
            break
    if log_file and os.path.exists(log_file):
        with open(log_file) as f:
            log_data = json.load(f)
            return log_data
    return None

# Regex for the folder format
folder_pattern = re.compile(r'anp_link_prediction_co_author_(.*?)_(\d{4}_\d{2}_\d{2})_\d{2}_\d{2}_\d{2}')

# Scan all the experiment folders
for experiment_folder in os.listdir(base_folder):
    experiment_path = os.path.join(base_folder, experiment_folder)
    if not os.path.isdir(experiment_path):
        continue

    # Check if the folder name matches the desired format and falls within the date range
    match = folder_pattern.match(experiment_folder)
    if not match:
        continue

    folder_date = match.group(2)
    if not (date_lower_int <= int(folder_date.replace("_", "")) <= date_upper_int):
        continue

    # Read info.json
    info = read_info_json(experiment_path)
    if info is None:
        continue
    
    # Read *_log.json
    log_data = read_log_json(experiment_path)
    if log_data is None:
        continue

    # Extract the values of interest
    prediction_type = info.get("only_new", "N/A")
    lr = info.get("lr", "N/A")

    if not (lr_lower <= float(lr) <= lr_upper):
        continue

   # Validation accuracy, highest accuracy, and average of last 10 epochs using the log file data
    validation_accuracy = log_data["validation_accuracy_list"][-1] if log_data["validation_accuracy_list"] else "N/A"
    highest_accuracy = max(log_data["validation_accuracy_list"]) if log_data["validation_accuracy_list"] else "N/A"
    avg_last_10_epochs = sum(log_data["validation_accuracy_list"][-10:]) / 10 if len(log_data["validation_accuracy_list"]) >= 10 else "N/A"
    num_epochs = len(log_data["validation_accuracy_list"])
    edge_number = info.get("edge_number", "N/A")
    aggregation_type = info.get("aggregation_type", "N/A")
    infosphere_parameters = info.get("infosphere_parameters", "N/A")
    infotype = info["infosphere_type"] if info.get("infosphere_type") else 0
    drop_percentage = info.get("drop_percentage", "N/A")
    if "hgt" in experiment_folder or "htg" in experiment_folder:
        network_type = "HGT"
    else:
        network_type = "SAGE + to_hetero"

    if not results[prediction_type][infotype].get(infosphere_parameters):
        results[prediction_type][infotype][infosphere_parameters] = []

    results[prediction_type][infotype][infosphere_parameters].append([
        experiment_folder,
        lr,
        validation_accuracy,
        highest_accuracy,
        avg_last_10_epochs,
        num_epochs,
        edge_number,
        aggregation_type,
        infotype,
        prediction_type,
        infosphere_parameters,
        drop_percentage,
        network_type
    ])

# Create a workbook and add worksheets
wb = Workbook()
for only_new in range(2):
    for i, entries2 in enumerate(results[only_new]):
        if entries2:
            sheet_name = f"{infosphere_string[i]}_{'only_new' if only_new else 'all_co_authors'}"
            ws = wb.create_sheet(title=sheet_name)

            # Add the header row
            headers = [
                "Folder", "Learning Rate", "Validation Accuracy", "Highest Accuracy", 
                "Average Last 10 Epochs", "Number of Epochs", "Edge Number", 
                "Aggregation Type", "Infosphere Type", "Only New", "Infosphere Parameters", "Drop Percentage", "Network Type"
            ]
            ws.append(headers)

            # Add data to worksheet
            for infosphere_parameter, entries in entries2.items():
                for entry in entries:
                    ws.append(entry)

# Remove the default sheet created
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save the workbook
wb.save(f"i_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

print("The Excel report was successfully created.")
